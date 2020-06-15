import csv
from datetime import datetime

from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import DeleteView, ListView, TemplateView, UpdateView, View


from openpyxl import Workbook

from rate.models import Rate
from rate.selectors import get_latest_rates
from rate.utils import display


class RateList(ListView):
    queryset = Rate.objects.all()
    template_name = 'list_all.html'


class RateDownloadCSV(View):
    HEADERS = (
        'id',
        'created',
        'source',
        'amount',
        'type',
    )
    queryset = Rate.objects.all().iterator()

    def get(self, request):
        # Create the HttpResponse object with the appropriate CSV header.
        response = self.get_response()

        writer = csv.writer(response)
        writer.writerow(self.__class__.HEADERS)

        for rate in self.queryset:
            values = []
            for attr in self.__class__.HEADERS:
                values.append(display(rate, attr))

            writer.writerow(values)

        return response

    def get_response(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rate.csv"'
        return response


class RateDownloadXLSX(View):
    HEADERS = (
        'id',
        'created',
        'source',
        'amount',
        'type',
    )
    queryset = Rate.objects.all().iterator()

    def get(self, request):

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        today = datetime.now().strftime("%Y-%m-%d")
        response['Content-Disposition'] = f'attachment; filename={today}-rates_all.xlsx'
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Rates_all'

        # Define the titles for columns
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.HEADERS, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for rate in self.queryset:
            row_num += 1
            # row
            row = []
            for attr in self.__class__.HEADERS:
                row.append(display(rate, attr))

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class LatestRatesView(TemplateView):
    template_name = 'latest-rates.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_list'] = get_latest_rates()
        return context


class UpdateRate(UserPassesTestMixin, UpdateView):
    template_name = 'rate-update.html'
    queryset = Rate.objects.all()
    fields = ('amount', 'source', 'currency_type', 'type')
    success_url = reverse_lazy('rate:list')

    def test_func(self):
        return self.request.user.is_authenticated and\
            self.request.user.is_superuser


class DeleteRate(UserPassesTestMixin, DeleteView):
    queryset = Rate.objects.all()
    success_url = reverse_lazy('rate:list')

    def get(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def test_func(self):
        return self.request.user.is_authenticated and\
            self.request.user.is_superuser
